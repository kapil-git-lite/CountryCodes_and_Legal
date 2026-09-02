# -*- coding: utf-8 -*-
import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from collections import defaultdict, Counter

# ---------- CONSTANTS (YOUR ORIGINAL LOGIC) ----------
MEMBERS_COL_NAME = "FAMILY MEMBERS"
STATUS_COL_NAME = "LEGAL STATUS"
PRIORITY = {"GRANTED": 1, "PENDING": 2, "LAPSED": 3, "REVOKED": 4, "EXPIRED": 5}

PRESENCE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
STATUS_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
DETAIL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

HEADER_FONT = Font(bold=True)
SUMMARY_FONT = Font(bold=True, italic=True)

def extract_country(member: str) -> str:
    m = member.strip().upper()
    if m.startswith("JPWO"):
        return "JP"
    return m[:2]

# ---------- CORE PROCESSING FUNCTION (MODIFIED FOR BYTES IO) ----------
def process_workbook(src_bytes: BytesIO, sheet_name: str = None) -> BytesIO:
    # Load workbook directly from the in-memory bytes
    wb = openpyxl.load_workbook(src_bytes)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Locate original columns from current header (Row 1)
    header_cells = ws[1]
    header_map = {cell.value: cell.column for cell in header_cells if cell.value}
    
    if MEMBERS_COL_NAME not in header_map or STATUS_COL_NAME not in header_map:
        raise ValueError(
            f"Could not find required columns '{MEMBERS_COL_NAME}' / "
            f"'{STATUS_COL_NAME}' in header row."
        )
    
    members_col = header_map[MEMBERS_COL_NAME]
    status_col = header_map[STATUS_COL_NAME]
    n_original_cols = ws.max_column
    n_original_rows = ws.max_row

    # Pass 1: Parse data rows
    row_data = []
    all_countries = set()

    for r in range(2, n_original_rows + 1):
        members_raw = ws.cell(r, members_col).value or ""
        status_raw = ws.cell(r, status_col).value or ""
        members = [m.strip() for m in str(members_raw).split("\n") if m.strip()]
        statuses = [s.strip().upper() for s in str(status_raw).split("\n") if s.strip()]

        per_country = defaultdict(Counter)
        for m, s in zip(members, statuses):
            cc = extract_country(m)
            per_country[cc][s] += 1
            all_countries.add(cc)

        row_data.append({"orig_row": r, "per_country": per_country})

    country_list = sorted(all_countries)

    # Insert summary row and shift data
    ws.insert_rows(1)
    data_start_row = 3
    data_end_row = n_original_rows + 1

    summary_label_cell = ws.cell(1, 1, "Filtered Count:")
    summary_label_cell.font = SUMMARY_FONT
    summary_label_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Build headers & formulas
    presence_start = n_original_cols + 1
    status_start = presence_start + len(country_list)
    detail_start = status_start + len(country_list)

    presence_cols = {}
    status_cols = {}
    detail_cols = {}

    for i, cc in enumerate(country_list):
        p_col = presence_start + i
        s_col = status_start + i
        d_col = detail_start + i
        presence_cols[cc] = p_col
        status_cols[cc] = s_col
        detail_cols[cc] = d_col

        # Headers on Row 2
        for col, fill, label in [(p_col, PRESENCE_FILL, cc), 
                                 (s_col, STATUS_FILL, f"{cc}_STATUS"), 
                                 (d_col, DETAIL_FILL, f"{cc}_DETAIL")]:
            cell = ws.cell(2, col, label)
            cell.font = HEADER_FONT
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

        # SUBTOTAL formulas on Row 1
        for col in [p_col, s_col, d_col]:
            col_letter = get_column_letter(col)
            cell = ws.cell(1, col, f"=SUBTOTAL(3, {col_letter}{data_start_row}:{col_letter}{data_end_row})")
            cell.font = SUMMARY_FONT
            cell.fill = SUMMARY_FILL
            cell.alignment = Alignment(horizontal="center")

    # Pass 2: Fill data rows
    for rec in row_data:
        r = rec["orig_row"] + 1
        per_country = rec["per_country"]

        for cc in country_list:
            if cc not in per_country:
                continue

            status_counts = per_country[cc]
            ws.cell(r, presence_cols[cc], "Y")
            best_status = min(status_counts.keys(), key=lambda s: PRIORITY.get(s, 99))
            ws.cell(r, status_cols[cc], best_status)
            ordered = sorted(status_counts.items(), key=lambda kv: PRIORITY.get(kv[0], 99))
            detail_str = ", ".join(f"{s}({n})" for s, n in ordered)
            ws.cell(r, detail_cols[cc], detail_str)

    # Cosmetics
    ws.freeze_panes = ws.cell(3, presence_start).coordinate
    for r in range(1, data_end_row + 1):
        ws.cell(r, members_col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, status_col).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions[get_column_letter(members_col)].width = 30
    ws.column_dimensions[get_column_letter(status_col)].width = 22
    for cc in country_list:
        ws.column_dimensions[get_column_letter(presence_cols[cc])].width = 6
        ws.column_dimensions[get_column_letter(status_cols[cc])].width = 12
        ws.column_dimensions[get_column_letter(detail_cols[cc])].width = 26

    # Save to in-memory BytesIO instead of a file path
    output_bytes = BytesIO()
    wb.save(output_bytes)
    output_bytes.seek(0)  # Rewind to the beginning so it can be read
    return output_bytes

# ---------- STREAMLIT UI ----------
st.set_page_config(page_title="Excel Legal Processor", layout="wide")
st.title("📊 Excel Country Codes & Legal Processor")
st.markdown("Upload your Excel file, and the system will extract country codes, aggregate legal statuses, and generate the processed output.")

uploaded_file = st.file_uploader("Choose an Excel file (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    # Show file details
    st.info(f"✅ Uploaded: **{uploaded_file.name}** (Size: {uploaded_file.size / 1024:.2f} KB)")

    # Process button (prevents auto-running on every UI refresh)
    if st.button("🚀 Process File"):
        try:
            # Show a spinner and progress bar for UX
            with st.spinner("Processing your file... This may take a moment."):
                # Call your processing function
                processed_bytes = process_workbook(uploaded_file)
            
            st.success("✅ Processing complete! Download your file below:")

            # Download button
            st.download_button(
                label="📥 Download Processed Excel",
                data=processed_bytes,
                file_name=f"processed_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"❌ An error occurred: {e}")
            st.stop()
